# -*- coding: UTF-8 -*-
import os, sys, subprocess
import pandas as pd
import numpy as np
from pathlib import Path
from tqdm import tqdm
import glob
import argparse
import shutil
import xlsxwriter
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
plt.style.use('ggplot')
import warnings
warnings.filterwarnings("ignore")


def smi2pred(ifilename, model_path, task_name_list,
            out_dir='tmp',
            shape=[331, 331, 3], layer_name='mixed10',
            grad_threshold=-1.0, is_gradcam_plus=False,
            is_TTA=False,
            run_gradcam_task_idx_list=None,
            smi2img_exe='/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/script/smi2img331x331'):
    """
    MOEでSMILESを画像に変換してpredict+GradCam
    Args:
        ifilename: SMILESファイルのパス
        model_path: モデルファイルのパス
        task_name_list: タスク名のリスト
        out_dir: ファイル出力先ディレクトリ
        shape: モデルの入力層のサイズ
        layer_name: GradCamで見るモデルの層名
        grad_threshold: GradCam実行するか決める予測スコアの閾値。デフォルトは-1.0として必ず実行
        is_gradcam_plus: GradCam++で実行するか。Falseだと普通のGradCam実行
        is_TTA: TTA入れるかどうか。FalseだとTTAなし
        run_gradcam_task_idx_list: GradCam実行するタスクidのリスト。Noneの時は全タスクGradCam実行する
        smi2img_exe: MOEのスクリプトのパス
    Return:
        予測結果のデータフレーム
    """
    # gpu確保されので関数内でimport
    import pathlib
    current_dir = pathlib.Path(__file__).resolve().parent # このファイルのディレクトリの絶対パスを取得
    sys.path.append( str(current_dir) + '/../' )#sys.path.append('/home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL')
    from predicter import grad_cam, multi_predict
    from transformer import get_train_valid_test

    import keras
    import keras.backend as K
    K.clear_session() # セッションのクリア
    K.set_learning_phase(0) # Test時には0にセット DropoutやBatchNormありmodelのときはロード前にこれがないとGradCamエラーになる
    model = keras.models.load_model(model_path, compile=False)

    os.makedirs(out_dir, exist_ok=True) # 出力ディレクトリ作成
    with open(ifilename, 'r') as ifs: # 入力ファイルロード
        ofilename_list = []
        smiles_list = []
        y_pred_list = []
        out_grad_cam_dir_list = []
        for i, line in tqdm(enumerate(ifs)):
            words = line.split()
            smiles = words[0]
            if len(words) == 2:
                # 入力ファイルの2列目にseries_no列ある場合
                series_no = words[1]
                ofilename = out_dir+'/%s.jpg' % (series_no)
            else:
                ofilename = out_dir+'/%07d.jpg' % (i)
            ofilename_list.append(ofilename)
            # smi2img
            try:
                smiles = max(smiles.split('.'), key=len) # LargestFragment にする
                subprocess.run([smi2img_exe, smiles, ofilename, str(0)]) # 外部プログラム同期処理
            except Exception as e:
                print('#### smi2img Exception smiles:', smiles, '#####')
                print(e)
                continue
            # predict
            try:
                X = get_train_valid_test.load_one_img(ofilename, shape[0], shape[1]) # 画像を1枚読み込んで、4次元テンソル(1, img_rows, img_cols, 3)へ変換+前処理
                if is_TTA == True:
                    # TTAやensembleする場合
                    y_pred = []
                    for x_i in list(X): # predict_tta()は1画像ずつであり、xは3次元ベクトルでないとだめ
                        y_pred_i = multi_predict.predict_tta(model, x_i
                                                             , TTA='flip'
                                                             , TTA_rotate_deg=0#30
                                                             , TTA_crop_num=0#, TTA_crop_size=[224, 224]
                                                             , preprocess=None, resize_size=None)
                        y_pred.append(y_pred_i)
                    y_pred = np.array(y_pred)
                else:
                    # 単純にpredict
                    y_pred = model.predict(X)
            except Exception as e:
                print('#### predict Exception smiles:', smiles, '#####')
                print(e)
                continue
            # gradcam
            try:
                out_grad_cam_dir = out_dir+'/'+Path(ofilename).stem
                os.makedirs(out_grad_cam_dir, exist_ok=True)
                img = keras.preprocessing.image.load_img(ofilename, target_size=shape[:2])
                x = keras.preprocessing.image.img_to_array(img)
                grad_cam_img = grad_cam.nobranch_multi_grad_cam(model,
                                                                out_grad_cam_dir,
                                                                Path(ofilename).name,
                                                                x,
                                                                None,
                                                                layer_name,
                                                                shape[0], shape[1],
                                                                grad_threshold=grad_threshold, is_gradcam_plus=is_gradcam_plus,
                                                                predicted_score=y_pred,
                                                                run_gradcam_task_idx_list=run_gradcam_task_idx_list
                                                                )
                for p in glob.glob(out_grad_cam_dir+"/**/*.jpg"):
                    shutil.copy2(p, out_grad_cam_dir) # GradCam画像コピー
                task_name_dir = [f for f in os.listdir(out_grad_cam_dir) if os.path.isdir(os.path.join(out_grad_cam_dir, f))]
                for d in task_name_dir:
                    shutil.rmtree(out_grad_cam_dir+'/'+d) # タスクごとのGradCam画像出力ディレクトリ削除
                shutil.copy2(ofilename, out_grad_cam_dir) # out_grad_cam_dirに元画像コピー
                os.remove(ofilename) # コピー元の元画像削除
            except Exception as e:
                print('#### gradcam Exception smiles:', smiles, '#####')
                print(e)
                continue
            # 全ての処理完了したらリストに詰める
            smiles_list.append(smiles)
            y_pred_list.append(np.round(y_pred, decimals=4))
            out_grad_cam_dir_list.append(out_grad_cam_dir)

    df_pred_score = None
    if len(y_pred_list) > 0:
        pred_score = np.array(y_pred_list).reshape(len(y_pred_list), len(task_name_list))
        cols = pd.MultiIndex.from_arrays([list(map(lambda i: 'task'+str(i), range(pred_score.shape[1]))), task_name_list]) # 2つのタスク名でヘッダ2行にする
        df_pred_score = pd.DataFrame(pred_score, columns=cols)

        # GradCam実行するタスクid列だけ残す
        if run_gradcam_task_idx_list is not None:
            run_gradcam_task_name_list = list(map(lambda t: 'task'+str(t), run_gradcam_task_idx_list))
            task_name_list = run_gradcam_task_name_list
            df_pred_score = df_pred_score.loc[:, task_name_list]

        df_pred_score['img_dir'] = out_grad_cam_dir_list
        df_pred_score['SMILES'] = smiles_list
        df_pred_score['id'] = list(map(lambda p: str(Path(p).stem), out_grad_cam_dir_list))
        df_pred_score = df_pred_score.set_index('id')
        df_pred_score = df_pred_score.iloc[:,[-1,-2]+list(range(0,len(task_name_list)))]

        # 化合物ごとに棒グラフ作成
        for index, series in df_pred_score.iterrows(): # 1行づつループ回す
            ## --- 棒グラフ ---
            #series[2:].plot.bar(alpha=0.6, figsize=(8*len(task_name_list)//15,8))#df_pred_score.iloc[index,2:].plot.bar(alpha=0.6, figsize=(8*len(task_name_list)//15,8))
            #plt.ylim([0,1])
            #plt.ylabel('predict score')
            ## ---------------
            # --- 横棒グラフ ---
            series_inv = series[series.index[::-1]] # seriesのindexの順番逆にする（plot.barhはindex逆順にしないとだめ）#df_T = df_pred_score.loc[:,df_pred_score.columns[::-1]].T
            series_inv[:-2].plot.barh(alpha=0.6, figsize=(8,8*len(task_name_list)//15), color='blue')#df_T.loc[:-2,index].T.plot.barh(alpha=0.6, figsize=(8,8*len(task_name_list)//15))
            plt.xlim([0,1])
            plt.xlabel('Active predict score')
            # -----------------
            plt.title(series['SMILES'].values[0], size=12)
            plt.grid(True) # グリッド線書く
            plt.savefig(series['img_dir'].values[0]+'/'+str(Path(series['img_dir'].values[0]).stem)+'_score_bar.jpg', bbox_inches="tight")
            plt.clf()

        # 全化合物についてスコアのヒストグラム作成
        df_pred_score.plot(bins=100, alpha=0.5, kind='hist')
        plt.xlim([0,1])
        plt.xlabel('Active predict score')
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0, fontsize=12) # 凡例を枠外に書く
        plt.title('All compounds', size=12)
        plt.grid(True) # グリッド線書く
        plt.savefig(out_dir+'/predict_hist.jpg', bbox_inches="tight")
        plt.clf()

    return df_pred_score

def cut_NPI_1class_GAR_2class(NPI_csv, GAR_csv):
    """
    NPI+GARのラベルデータの列減らす
    Args:
        csv: train/validation/test set列をつけたNPIのラベル情報csv
    """
    NPI_task_name_list = cut_1class(NPI_csv)
    GAR_task_name_list = cut_2class(GAR_csv)
    NPI_task_name_list.extend(GAR_task_name_list)
    return NPI_task_name_list

def cut_2class(csv):
    """
    <=0.01um と <=0.001um のラベルデータの列減らす
    Args:
        csv: train/validation/test set列をつけたラベル情報csv
    """
    df_label = pd.read_csv(csv)
    # 不要なラベル列削除
    drop_task_name_list = []
    for task_name in list(df_label.columns[4:-2]):
        if '<=0.01um' in task_name:
            drop_task_name_list.append(task_name)
        if '<=0.001um' in task_name:
            drop_task_name_list.append(task_name)
    df_label = df_label.drop(drop_task_name_list, axis=1)
    task_name_list = list(df_label.columns)[4:-2]
    return task_name_list

def cut_1class(csv):
    """
    <=0.001um のラベルデータの列減らす
    Args:
        csv: train/validation/test set列をつけたNPIのラベル情報csv
    """
    df_label = pd.read_csv(csv)
    # 不要なラベル列削除
    drop_task_name_list = []
    for task_name in list(df_label.columns[4:-2]):
        if '<=0.001um' in task_name:
            drop_task_name_list.append(task_name)
    df_label = df_label.drop(drop_task_name_list, axis=1)
    task_name_list = list(df_label.columns)[4:-2]
    return task_name_list

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--output_dir", type=str, required=True,
                    help="output dir path.")
    ap.add_argument("-s", "--smiles_path", type=str, required=True,
                    help="SMILES file path.")
    ap.add_argument("-mt", "--model_type", type=str, default='class',
                    help="model type (ex. 'class').")#, 'regression'")
    ap.add_argument("-t", "--thema", type=str, required=True, choices=['NPI', 'GAR', 'kinase', 'NPI_GAR'],#['NPI', 'GAR', 'kinase', 'NPI_GAR', 'cut_NPI_1class', 'cut_GAR_2class', 'cut_NPI_1class_GAR_2class'],
                    help="theme codes (ex. 'NPI', 'GAR', 'kinase', 'NPI_GAR').")
    ap.add_argument("-mc", "--method_category", type=str, default='Pharmacology',
                    help="method category names.")
    ap.add_argument("-mk", "--method_kind", type=str, default='IC50',
                    help="method kind names.")
    ap.add_argument("-gt", "--grad_threshold", type=float, default=0.5,
                    help="threshold to execute Grad-Cam.")
    ap.add_argument("-gp", "--is_gradcam_plus", action='store_const', const=True, default=False,
                    help="Grad-Cam++ flag.")
    ap.add_argument("-tta", "--is_TTA", action='store_const', const=True, default=False,
                    help="TTA(flip) flag.")
    ap.add_argument("-cut_NPI", "--is_cut_NPI", action='store_const', const=True, default=False,
                    help="flag that displays only GAR when NPI + GAR.")
    ap.add_argument("-m", "--model_path", type=str,
                    help="used to specify model file path (ex. /home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL/experiment/output_class_kinase_result_IC50_scaffold/best_model_val_acc.h5).")
    ap.add_argument("-ln", "--layer_name", type=str,
                    help="used to specify Grad-Cam CNN layer name (ex. 'mixed10').")
    args = vars(ap.parse_args())

    run_gradcam_task_idx_list = None # Grad-Cam実行するtaskid記録用変数。Noneなら全taskでGrad-Cam実行
    is_sparkline = True # エクセルのスパークラインむだだから描画するか

    if args['model_type'] == 'class':
        #print(args['model_type'], args['thema'], args['method_category'], args['method_kind'])
        ####------------------ best modelではなかった分類taskのクラスcutしないNPI、GAR、NPI_GARの条件 ------------------###
        #if args['thema'] == 'NPI' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
        #    data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
        #    task_name_list = list(pd.read_csv(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv').columns[4:-2])
        #    assay_num = 3
        #    class_num = 5
#
        #if args['thema'] == 'GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
        #    data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
        #    task_name_list = list(pd.read_csv(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv').columns[4:-2])
        #    assay_num = 16
        #    class_num = 5
#
        #if args['thema'] == 'NPI_GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
        #    data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
        #    NPI_task_name_list = list(pd.read_csv(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv').columns[4:-2])
        #    GAR_task_name_list = list(pd.read_csv(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv').columns[4:-2])
        #    NPI_task_name_list.extend(GAR_task_name_list)
        #    task_name_list = NPI_task_name_list
        #    assay_num = 3+16
        #    class_num = 5
        #    # NPI+GARからNPIタスクを除く場合
        #    if args['is_cut_NPI'] == True:
        #        assay_num = 16
        #        class_num = 5
        #        GAR_task_name_list = list(pd.read_csv(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv').columns[4:-2])
        #        NPI_task_name_list = list(pd.read_csv(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv').columns[4:-2])
        #        run_gradcam_task_idx_list = list(range(len(NPI_task_name_list), len(task_name_list)))

        ###------------------ best modelのpathハードコードしている分類taskのクラスcutするNPI、GAR、NPI_GARの条件 ------------------###
        if args['thema'] == 'NPI' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
            data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
            task_name_list = cut_1class(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv')
            assay_num = 3
            class_num = 5-1
            # cut_NPI_1class best model(Xception)
            model_path = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL/tuning/output_optuna_class_NPI_cut_task.Pharmacology.represent_IC50_scaffold_roc/best_trial_loss.h5'
            layer_name = 'block14_sepconv2_act'

        if args['thema'] == 'GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
            data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
            task_name_list = cut_2class(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv')
            assay_num = 16
            class_num = 5-2
            # cut_GAR_2class best model(InceptionV3)
            model_path = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL/experiment/output_class_GAR.Pharmacology.represent_IC50_scaffold_3task/best_model.h5'
            layer_name = 'mixed10'

        if args['thema'] == 'NPI_GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
            data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess/data/dataset/'
            task_name_list = cut_NPI_1class_GAR_2class(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv',
                                                       data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv')
            assay_num = 3+16
            class_num = 4+3
            # NPI+GARからNPIタスクを除く場合
            if args['is_cut_NPI'] == True:
                assay_num = 16
                class_num = 3
                GAR_task_name_list = cut_2class(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/class/GAR.Pharmacology.csv')
                NPI_task_name_list = cut_1class(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/class/NPI.Pharmacology.csv')
                run_gradcam_task_idx_list = list(range(len(NPI_task_name_list), len(task_name_list)))
            # cut_NPI_1class_GAR_2class best(InceptionV3) NPI,GAR単独モデルの方がmeanAUC若干高い
            model_path = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL/tuning/output_optuna_class_NPI_GAR_cut_task.Pharmacology.represent_IC50_scaffold_roc/best_trial_loss.h5'
            layer_name = 'mixed10'

        if args['thema'] == 'kinase' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
            data_dir = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/preprocess_v2/data/dataset/'
            task_name_list = list(pd.read_csv(data_dir+'kinase_result/scaffold/class/kinase_result.csv').columns[4:-2])
            assay_num = 61
            class_num = 1
            # kinase_result best(InceptionV3)
            model_path = '/home/aaa00162/jupyterhub/notebook/H3-050/Submit/DL/experiment/output_class_kinase_result_IC50_scaffold/best_model.h5'
            layer_name = 'mixed10'
            is_sparkline = False # kinase_resultはタスクごとのクラスは1つ(<=10uM)だけ。エクセルのスパークラインむだだから描画させない

    #if args['model_type'] == 'regression':
    #    ###------------------ 回帰taskのNPI、GAR、NPI_GARの条件 ------------------###
    #    if args['thema'] == 'NPI' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
    #        task_name_list = list(pd.read_csv(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/regression/NPI.Pharmacology.csv').columns[4:-2])
    #        task_name_list = [n.replace('.1', '') for n in task_name_list] # 回帰のcsvは列名に「.1」がついてるので消す
    #        assay_num = 3
    #        class_num = 5
    #
    #    if args['thema'] == 'GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
    #        task_name_list = list(pd.read_csv(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/regression/GAR.Pharmacology.csv').columns[4:-2])
    #        task_name_list = [n.replace('.1', '') for n in task_name_list] # 回帰のcsvは列名に「.1」がついてるので消す
    #        assay_num = 16
    #        class_num = 5
    #
    #    if args['thema'] == 'kinase' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
    #        task_name_list = list(pd.read_csv(data_dir+'kinase/scaffold/regression/kinase.csv').columns[4:-2])
    #        task_name_list = [n.replace('.1', '') for n in task_name_list] # 回帰のcsvは列名に「.1」がついてるので消す
    #        assay_num = 60
    #        class_num = 5
    #
    #    if args['thema'] == 'NPI_GAR' and args['method_category'] == 'Pharmacology' and args['method_kind'] == 'IC50':
    #        NPI_task_name_list = list(pd.read_csv(data_dir+'NPI.Pharmacology.represent_IC50/scaffold/regression/NPI.Pharmacology.csv').columns[4:-2])
    #        GAR_task_name_list = list(pd.read_csv(data_dir+'GAR.Pharmacology.represent_IC50/scaffold/regression/GAR.Pharmacology.csv').columns[4:-2])
    #        NPI_task_name_list.extend(GAR_task_name_list)
    #        task_name_list = NPI_task_name_list
    #        task_name_list = [n.replace('.1', '') for n in task_name_list] # 回帰のcsvは列名に「.1」がついてるので消す
    #        assay_num = 3+16
    #        class_num = 5

    # 引数で指定したモデル使う場合
    if args['model_path'] is not None:
        model_path = args['model_path']
        print('model_path:', model_path)
    if args['layer_name'] is not None:
        layer_name = args['layer_name']
        print('layer_name:', layer_name)

    df_pred_score = smi2pred(args['smiles_path'],
                            model_path,
                            task_name_list,
                            out_dir=args['output_dir'],
                            shape=[331, 331, 3], layer_name=layer_name,
                            grad_threshold=args['grad_threshold'], is_gradcam_plus=args['is_gradcam_plus'],
                            is_TTA=args['is_TTA'],
                            run_gradcam_task_idx_list=run_gradcam_task_idx_list
                            )
    # 予測結果0件の時
    if df_pred_score is None:
        print('#### predict result None. ####')
        return

    # NPI+GARからNPIタスクを除く場合
    if args['is_cut_NPI'] == True and (args['thema'] == 'cut_NPI_1class_GAR_2class' or args['thema'] == 'NPI_GAR'):
        task_name_list = GAR_task_name_list

    # ---------------------- エクセル編集 ---------------------- #
    #writer = pd.ExcelWriter(Path(args['smiles_path']).stem+'_predict.xlsx', engine='xlsxwriter') # カレントディレクトリに予測結果エクセル出力
    writer = pd.ExcelWriter(args['output_dir']+'/'+str(Path(args['smiles_path']).stem)+'_predict.xlsx', engine='xlsxwriter') # 出力先dirに予測結果エクセル出力
    df_pred_score.to_excel(writer, sheet_name='predict') # MultiIndexでは index=False 使えない
    workbook  = writer.book
    worksheet = writer.sheets['predict']
    header_num = 3 # ヘッダ行の数
    index_col_num = 3 # score列以外のindexに相当する列数
    score_col_start_num = 4 # score列開始列番号
    # 最終列のアルファベット取得
    wb = openpyxl.Workbook()
    ws = wb.active
    coordinate = ws.cell(row=1,column=len(task_name_list)+index_col_num).coordinate
    max_col = coordinate.replace("1", "")
    # 条件付き書式で数値セルの文字色変更
    #cell_red_format = workbook.add_format()
    #cell_red_format.set_pattern(1)  # This is optional when using a solid fill.
    #cell_red_format.set_bg_color('red')
    worksheet.conditional_format('D4:'+max_col+str(df_pred_score.shape[0]+header_num),
                                    {'type':'cell',
                                     'criteria':'>=','value':0.5, # 0.5以上なら条件適用
                                     'format':workbook.add_format({'font_color':'red'}) # 文字色赤にする
                                     #'format':cell_red_format # 文字色赤にする データバーあると見にくいからやめる
                                     }
                                )
    # 条件付き書式のデータバー 全scoreの範囲について
    worksheet.conditional_format('D4:'+max_col+str(df_pred_score.shape[0]+header_num), {'type': 'data_bar'})

    # テーマごとのタスク数一致するときだけスパークラインの棒グラフ列作る
    if assay_num*class_num == len(task_name_list) and is_sparkline == True:
        # スパークラインの棒グラフ
        ##worksheet.add_sparkline(max_col_plus+str(i+5), {'range':'predict!'+score_col+str(i+4)+':'+max_col+str(i+4),'type':'column'}) # 1化合物について書く場合
        for i in range(df_pred_score.shape[0]):
            score_col_start = ws.cell(row=1,column=i+score_col_start_num).coordinate.replace("1", "")
            # 各assayごとに棒グラフ作成
            for a_n in range(assay_num):
                assay_start_col = ws.cell(row=1,column=(len(task_name_list)//assay_num)*a_n + score_col_start_num).coordinate.replace("1", "")
                assay_end_col   = ws.cell(row=1,column=(len(task_name_list)//assay_num)*(a_n+1) + score_col_start_num - 1).coordinate.replace("1", "")
                max_col_plus = ws.cell(row=1,column=len(task_name_list)+score_col_start_num+a_n).coordinate.replace("1", "") # 最終列+1のアルファベット取得

                # NPI+GARからNPIタスクを除く場合
                if args['is_cut_NPI'] == True and (args['thema'] == 'cut_NPI_1class_GAR_2class' or args['thema'] == 'NPI_GAR'):
                    col_name = 'task'+str(len(NPI_task_name_list)+len(task_name_list)//assay_num*a_n)+'-'+str(len(NPI_task_name_list)+len(task_name_list)//assay_num*(a_n+1)-1)
                else:
                    col_name = 'task'+str(len(task_name_list)//assay_num*a_n)+'-'+str(len(task_name_list)//assay_num*(a_n+1)-1)

                worksheet.write(max_col_plus+'1', col_name, workbook.add_format({'bold': True}))
                worksheet.add_sparkline(max_col_plus+str(i+header_num+1),
                                        {'range':'predict!'+assay_start_col+str(i+header_num+1)+':'+assay_end_col+str(i+header_num+1),
                                         'type':'column'})
                # 条件付き書式のデータバー 各assayごとに作る場合
                #worksheet.conditional_format(assay_start_col+str(i+header_num+1)+':'+assay_end_col+str(i+header_num+1), {'type': 'data_bar'})

    # ハイパーリンク
    for i,dir in enumerate(df_pred_score['img_dir']):
        worksheet.write_url('C'+str(i+header_num+1), str(Path(dir).stem))#dir)
    # オートフィルタ
    worksheet.autofilter('A3:'+max_col+str(df_pred_score.shape[1]+header_num))
    # 全化合物についてスコアのヒストグラム画像挿入
    worksheet2 = workbook.add_worksheet('hist')
    worksheet2.insert_image('A1', args['output_dir']+'/predict_hist.jpg')
    # 保存
    workbook.close()
    writer.save()
    # --------------------------------------------------------- #

if __name__ == '__main__':
    main()
